import os
import re
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
from pathlib import Path
from src import config

# path for excel file
EXCEL_PATH = config.EXCEL_PATH

# folder transcripts file
BASE_DIR = Path(__file__).parent.parent
TRANSCRIPTS_DIR = BASE_DIR / "files"

def analyze_transcript(text: str) -> dict:
    """
        analyze transcript and create row data for Excel
    """
    result = {
        "Дата": datetime.datetime.now().strftime("%Y-%m-%d"),
        "Тип звернення": "Сервіс",
        "Номер телефону": "пропускаємо",
        "Філія": "пропускаємо",
        "Менеджер": "",
        "Початок розмови, представлення": "ні",
        "Чи дізнався менеджер кузов автомобіля": "ні",
        "Чи дізнався менеджер рік автомобіля": "ні",
        "Чи дізнався менеджер пробіг": "ні",
        "Пропозиція про комплексну діагностику": "ні",
        "Дізнався які роботи робилися раніше": "ні",
        "Запис на сервіс, Дата": "ні",
        "Завершення розмови прощання": "ні",
        "Яка робота з топ 100": "",
        "Чи дотримувався всіх інструкцій з топ 100 робіт Да/Ні": "ні",
        "Яких рекомендацій менеджер не дотримувався з топ 100 робіт": "пропускаємо",
        "Результат": 0,
        "Дата 2": "",
        "Тип Звернення 2": "",
        "Запчастини": "",
        "Коментар": "",
    }

    text = text.lower()

    # keywords for type of call
    if "то" in text:
        result["Тип звернення"] = "ТО"
    elif "діагностик" in text or "перевірк" in text:
        result["Тип звернення"] = "Діагностика"
    elif "консультац" in text:
        result["Тип звернення"] = "Консультація"

    # keywords
    if re.search(r"добр(ий|ого) день|дня|вітаю", text):
        result["Початок розмови, представлення"] = "так"
    if re.search(r"кузов|седан|універсал|хетчбек|купе", text):
        result["Чи дізнався менеджер кузов автомобіля"] = "так"
    if re.search(r"(\b19\d{2}\b|\b20\d{2}\b)", text):
        result["Чи дізнався менеджер рік автомобіля"] = "так"
    if "пробіг" in text or "кілометр" in text:
        result["Чи дізнався менеджер пробіг"] = "так"
    if "діагностик" in text or "перевірк" in text:
        result["Пропозиція про комплексну діагностику"] = "так"
    if "робот" in text and "раніше" in text:
        result["Дізнався які роботи робилися раніше"] = "так"
    if "запишу" in text or "записати" in text or "сервіс" in text:
        result["Запис на сервіс, Дата"] = "так"
    if re.search(r"до побачення|гарного дня|дякую", text):
        result["Завершення розмови прощання"] = "так"

    # top 100 works
    found_work = None
    for work in config.TOP_WORKS:
        if work.lower() in text:
            found_work = work
            break

    result["Яка робота з топ 100"] = found_work if found_work else ""
    result["Чи дотримувався всіх інструкцій з топ 100 робіт Да/Ні"] = "так" if found_work else "ні"
    result["Менеджер"] = get_manager_name(text)

    # comments
    comments = []
    for key, desc in {
        "Початок розмови, представлення": "не привітався/не представився",
        "Чи дізнався менеджер кузов автомобіля": "не дізнався кузов автомобіля",
        "Чи дізнався менеджер рік автомобіля": "не дізнався рік автомобіля",
        "Чи дізнався менеджер пробіг": "не дізнався пробіг",
        "Пропозиція про комплексну діагностику": "не запропонував діагностику",
        "Дізнався які роботи робилися раніше": "не дізнався робіт раніше",
        "Запис на сервіс, Дата": "не зробив запис на сервіс",
        "Завершення розмови прощання": "не попрощався з клієнтом",
    }.items():
        if result[key] == "ні":
            comments.append(f"не ок: {desc}")

    result["Коментар"] = "\n".join(comments) if comments else "Все виконано коректно"

    # counting result
    result["Результат загальний"] = 1 if all(result[key] == "так" for key in [
        "Початок розмови, представлення",
        "Чи дізнався менеджер кузов автомобіля",
        "Чи дізнався менеджер рік автомобіля",
        "Чи дізнався менеджер пробіг",
        "Пропозиція про комплексну діагностику",
        "Дізнався які роботи робилися раніше",
        "Запис на сервіс, Дата",
        "Завершення розмови прощання"
    ]) else 0

    actions = [
        "Початок розмови, представлення",
        "Чи дізнався менеджер кузов автомобіля",
        "Чи дізнався менеджер рік автомобіля",
        "Чи дізнався менеджер пробіг",
        "Пропозиція про комплексну діагностику",
        "Дізнався які роботи робилися раніше",
        "Запис на сервіс, Дата",
        "Завершення розмови прощання"
    ]

    result["Результат"] = 1 + round(sum(result[key]=="так" for key in actions) * (9/len(actions)))

    return result


def write_to_excel(row_data: dict):
    """
        added row to Excel
    """
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    start_row = ws.max_row + 1
    for col_idx, key in enumerate(row_data.keys(), start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=row_data[key])
        if key == "Коментар" and "не ок" in row_data[key]:
            cell.font = Font(color="FF0000")

    wb.save(EXCEL_PATH)
    wb.close()


def process_all_transcripts():
    """
        read all transcripts and analyze them
    """
    files = [f for f in os.listdir(TRANSCRIPTS_DIR) if f.endswith(".txt")]

    if not files:
        print("no transcripts found")
        return

    print(f"Find {len(files)} files to analyze.")

    for idx, file in enumerate(files, start=1):
        path = os.path.join(TRANSCRIPTS_DIR, file)

        with open(path, "r", encoding="utf-8") as f:
            text = f.read()

        print(f"[{idx}/{len(files)}] analyze {file} ...")

        row_data = analyze_transcript(text)
        row_data["Номер телефону"] = "пропускаємо"
        write_to_excel(row_data)

    print("all transcripts added to Excel")


def get_manager_name(text: str) -> str:
    text = text.lower()

    # template for "менеджер імʼя"
    match = re.search(r"менеджер\s+([А-Яа-яЁёЇїІіЄєҐґ\w-]+)", text)
    if match:
        return match.group(1).capitalize()

    # template for "я імʼя"
    match = re.search(r"\bя\s+([А-Яа-яЁёЇїІіЄєҐґ\w-]+)", text)
    if match:
        return match.group(1).capitalize()

    return "пропускаємо"


if __name__ == "__main__":
    process_all_transcripts()
